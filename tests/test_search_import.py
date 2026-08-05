"""Tests for global search and Excel bulk import."""

from datetime import date

from openpyxl import Workbook, load_workbook

from timber.core.excel_import import (
    PARTY_HEADERS,
    import_parties,
    import_transactions,
    write_party_template,
)
from timber.core.payment_service import create_payment
from timber.core.search import global_search
from timber.core.transaction_service import create_bapari_txn
from timber.db.models import BapariTxn, Party, WoodType
from timber.db.models.party import PARTY_BAPARI


# --- search ----------------------------------------------------------
def _bapari(session, name="Karim", phone="0300"):
    from timber.core import admin_service

    return admin_service.create_party(
        session, name=name, party_type=PARTY_BAPARI, phones=[phone]
    )


def test_search_empty(session):
    assert global_search(session, "") == []


def test_search_by_party_and_vehicle(session):
    p = _bapari(session)
    create_bapari_txn(
        session, txn_date=date(2026, 1, 1), party_id=p.id, weight=10, rate=1000,
        vehicle_no="LEA-1",
    )
    create_payment(session, txn_date=date(2026, 1, 2), party_id=p.id, amount=500)

    by_name = global_search(session, "Karim")
    kinds = {r.kind for r in by_name}
    assert "party" in kinds and "purchase" in kinds and "payment" in kinds

    by_vehicle = global_search(session, "LEA-1")
    assert any(r.kind == "purchase" for r in by_vehicle)


# --- import ----------------------------------------------------------
def test_party_template_headers(tmp_path):
    path = write_party_template(tmp_path / "t.xlsx")
    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == PARTY_HEADERS


def _make_file(tmp_path, name, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    path = tmp_path / name
    wb.save(path)
    return path


def test_import_parties(session, tmp_path):
    path = _make_file(
        tmp_path, "p.xlsx", PARTY_HEADERS,
        [
            ["Karim", "bapari", "0300", "", "", "Lahore", 0, ""],
            ["ABC", "factory", "", "", "", "Lahore", 1000, 30],
            ["Bad", "wholesaler", "", "", "", "", 0, ""],  # invalid type
        ],
    )
    report = import_parties(session, path)
    assert report.created == 2
    assert len(report.errors) == 1
    assert report.errors[0][0] == 4  # row number
    assert session.query(Party).count() == 2


def test_import_transactions(session, tmp_path):
    _bapari(session, name="Karim")
    header = [
        "date", "party", "vehicle", "wood_type", "location",
        "weight", "rate", "freight", "loading", "notes",
    ]
    path = _make_file(
        tmp_path, "t.xlsx", header,
        [
            ["2026-01-01", "Karim", "LEA-1", "Kikar", "Lahore", 10, 1000, 0, 0, ""],
            ["2026-01-02", "Ghost", "LEA-2", "Kikar", "Lahore", 5, 1000, 0, 0, ""],
        ],
    )
    report = import_transactions(session, path, PARTY_BAPARI)
    assert report.created == 1
    assert len(report.errors) == 1
    assert session.query(BapariTxn).count() == 1
    assert session.query(WoodType).filter_by(name="Kikar").count() == 1
