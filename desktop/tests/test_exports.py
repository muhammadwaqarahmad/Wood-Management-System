"""Export formats: PDF (incl. Urdu font) and Excel (real numbers)."""

from openpyxl import load_workbook

from timber import i18n
from timber.core import excel_export, pdf_export
from timber.core.report_data import ReportData


def _sample():
    return ReportData(
        title="بیوپاری کھاتہ",
        headers=["تاریخ", "گاڑی", "بل", "کرایہ"],
        rows=[
            ["2026-06-26", "LEA-1", "10,500.00", "500.00 (کریم)"],
            ["2026-06-27", "GLT-9", "4,000.00", ""],
        ],
        summary=[("کل خریداری", "14,500.00")],
    )


def test_pdf_writes_for_both_languages(tmp_path):
    for lang in ("en", "ur"):
        i18n.set_language(lang)
        p = tmp_path / f"r_{lang}.pdf"
        pdf_export.write(_sample(), p)
        assert p.exists() and p.stat().st_size > 1000


def test_urdu_font_has_no_missing_glyphs():
    from reportlab.pdfbase import pdfmetrics

    from timber.core.pdf_fonts import register_fonts, shape
    name = register_fonts()
    assert name, "an Urdu/Arabic font must be bundled"
    face = pdfmetrics.getFont(name).face
    for word in ["بیوپاری", "کھاتہ", "کرایہ", "عبدالستار", "ووڈز", "ادا"]:
        for ch in shape(word):
            if ch.strip():
                assert face.charToGlyph.get(ord(ch), 0) != 0, f"missing glyph in {word!r}"


def test_excel_numbers_are_real(tmp_path):
    i18n.set_language("en")
    p = tmp_path / "r.xlsx"
    excel_export.write(_sample(), p)
    wb = load_workbook(p)
    ws = wb.active
    found = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, (int, float))]
    assert 10500.0 in found and 14500.0 in found  # parsed, not strings
