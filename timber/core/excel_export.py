"""Render a ReportData to a polished .xlsx using openpyxl.

Branded title block, styled header, thin borders, frozen header row,
right-aligned real numbers (so totals/sorting work), a bold summary
block, auto column widths, and RTL sheet direction for Urdu.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from timber import config, i18n
from timber.core.report_data import ReportData

_BRAND = "1565C0"
_BRAND_DARK = "0F3D7A"
_ALT = "EEF3FB"
_NUM_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")
_THIN = Side(style="thin", color="C7D0DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_THICK = Side(style="medium", color="0F172A")  # bold dark divider line


def _num(value):
    """Return a float if the cell is a plain number string, else None."""
    if not isinstance(value, str):
        return None
    s = value.strip()
    if _NUM_RE.match(s):
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def write(report: ReportData, path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    if i18n.get_language() == "ur":
        ws.sheet_view.rightToLeft = True

    ncols = max(len(report.headers), 1)
    last_col = get_column_letter(ncols)

    # --- title block ---
    name = config.APP_NAME_UR if i18n.get_language() == "ur" else config.APP_NAME
    ws.append([name])
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].font = Font(bold=True, size=16, color=_BRAND_DARK)
    ws.append([report.title])
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].font = Font(bold=True, size=12)
    ws.append([f"{i18n.tr('generated_on')}: {datetime.now():%Y-%m-%d %H:%M}"])
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws.append([])

    def _write_cell(r: int, c: int, value, *, bold=False, fill=None):
        cell = ws.cell(row=r, column=c)
        num = _num(value)
        if num is not None:
            cell.value = num
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.value = value
        cell.border = _BORDER
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        return cell

    def _write_table(headers, rows, bold_rows):
        hr = ws.max_row + 1
        ws.append(headers)
        for cell in ws[hr]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_BRAND)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = _BORDER
        for i, row in enumerate(rows):
            r = ws.max_row + 1
            for c, value in enumerate(row, start=1):
                _write_cell(
                    r, c, value,
                    bold=i in bold_rows,
                    fill="E2E8F0" if i in bold_rows else (_ALT if i % 2 else None),
                )
        # Bold dark divider line: a thick right border on the divider column.
        if report.divider_after is not None:
            dcol = report.divider_after + 1  # 1-indexed; line on its right edge
            if 1 <= dcol <= len(headers):
                for rr in range(hr, ws.max_row + 1):
                    cell = ws.cell(row=rr, column=dcol)
                    b = cell.border
                    cell.border = Border(left=b.left, right=_THICK, top=b.top, bottom=b.bottom)
        return hr

    # --- hero headline (e.g. Total business worth) ---
    if report.hero:
        label, value = report.hero
        r = ws.max_row + 1
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        vcell = _write_cell(r, 2, value)
        vcell.font = Font(bold=True, color="FFFFFF", size=14)
        vcell.fill = PatternFill("solid", fgColor="4F46E5")
        ws.append([])

    # --- stat tiles (overall numbers) as label/value pairs ---
    if report.tiles:
        for label, value in report.tiles:
            r = ws.max_row + 1
            _write_cell(r, 1, label, bold=True, fill=_ALT)
            _write_cell(r, 2, value, bold=True, fill=_ALT)
        ws.append([])

    header_row = ws.max_row + 1
    if report.sections:
        # --- titled sections, each its own table ---
        for sec in report.sections:
            r = ws.max_row + 1
            tcell = ws.cell(row=r, column=1, value=sec.title)
            tcell.font = Font(bold=True, size=12, color=_BRAND_DARK)
            _write_table(sec.headers, sec.rows, set(sec.bold_rows))
            ws.append([])
    elif report.headers or report.rows:
        # --- classic single data table ---
        header_row = _write_table(report.headers, report.rows, set())

    # --- summary ---
    if report.summary:
        ws.append([])
        for label, value in report.summary:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            num = _num(value)
            vcell = ws.cell(row=r, column=2)
            if num is not None:
                vcell.value = num
                vcell.number_format = "#,##0.00"
            else:
                vcell.value = value
            vcell.font = Font(bold=True)

    # Freeze the header row (single-table reports only) and size columns.
    if not report.sections:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(width, 45)

    wb.save(str(path))
    return path
