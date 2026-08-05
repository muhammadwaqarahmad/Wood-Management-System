"""Bulk import of parties and transactions from .xlsx files.

Each data row is imported inside its own SAVEPOINT, so a single bad row
is reported and skipped without aborting the rest. The caller commits.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from timber.core import admin_service
from timber.core.transaction_service import create_bapari_txn, create_factory_txn
from timber.db.models import Location, Party, WoodType
from timber.db.models.party import PARTY_BAPARI, PARTY_FACTORY


def _name_matches(name: str):
    """Match a party by name in either language (or the physical fallback)."""
    return or_(Party.name_en == name, Party.name_ur == name, Party._name == name)

PARTY_HEADERS = [
    "name", "type", "phone", "email", "address",
    "location", "opening_balance", "credit_days",
]
TXN_HEADERS = [
    "date", "party", "vehicle", "wood_type", "location",
    "weight", "rate", "freight", "loading", "notes",
]


@dataclass
class ImportReport:
    created: int = 0
    errors: list[tuple[int, str]] = field(default_factory=list)


def _str(value) -> str:
    return "" if value is None else str(value).strip()


def _parse_date(value) -> date:
    if value in (None, ""):
        raise ValueError("date is required")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(_str(value), "%Y-%m-%d").date()


def _rows(path: str | Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return []
    headers = [_str(h).lower() for h in raw[0]]
    out = []
    for row in raw[1:]:
        if row is None or all(c is None for c in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def _get_or_create_lookup(session: Session, model: type, name: str) -> int | None:
    name = _str(name)
    if not name:
        return None
    obj = session.scalar(select(model).where(model.name == name))
    if obj is None:
        obj = model(name=name)
        session.add(obj)
        session.flush()
    return obj.id


# --- templates -------------------------------------------------------
def write_party_template(path: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(PARTY_HEADERS)
    ws.append(
        ["Karim Timber", "bapari", "0300-1234567", "karim@example.com",
         "Main Bazaar", "Lahore", 0, ""]
    )
    wb.save(str(path))
    return Path(path)


def write_txn_template(path: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(TXN_HEADERS)
    ws.append(
        ["2026-01-01", "Karim Timber", "LEA-1234", "Kikar", "Lahore", 10, 1000, 500, 200, ""]
    )
    wb.save(str(path))
    return Path(path)


# --- importers -------------------------------------------------------
def import_parties(session: Session, path: str | Path, created_by=None) -> ImportReport:
    report = ImportReport()
    for i, row in enumerate(_rows(path), start=2):
        try:
            with session.begin_nested():
                name = _str(row.get("name"))
                ptype = _str(row.get("type")).lower()
                if ptype not in (PARTY_BAPARI, PARTY_FACTORY):
                    raise ValueError("type must be 'bapari' or 'factory'")
                if session.scalar(
                    select(Party).where(_name_matches(name), Party.party_type == ptype)
                ):
                    raise ValueError("already exists")
                loc_id = _get_or_create_lookup(session, Location, row.get("location"))
                phone = _str(row.get("phone"))
                credit_raw = row.get("credit_days")
                admin_service.create_party(
                    session,
                    name=name,
                    party_type=ptype,
                    email=_str(row.get("email")),
                    address=_str(row.get("address")),
                    credit_days=int(credit_raw) if credit_raw else None,
                    phones=[phone] if phone else None,
                    location_id=loc_id,
                    opening_balance=row.get("opening_balance") or 0,
                    created_by=created_by,
                )
            report.created += 1
        except Exception as exc:  # noqa: BLE001
            report.errors.append((i, str(exc)))
    return report


def import_transactions(
    session: Session, path: str | Path, side: str, created_by=None
) -> ImportReport:
    create = create_bapari_txn if side == PARTY_BAPARI else create_factory_txn
    report = ImportReport()
    for i, row in enumerate(_rows(path), start=2):
        try:
            with session.begin_nested():
                party_name = _str(row.get("party"))
                party = session.scalar(
                    select(Party).where(
                        _name_matches(party_name), Party.party_type == side
                    )
                )
                if party is None:
                    raise ValueError(f"{side} '{party_name}' not found")
                create(
                    session,
                    txn_date=_parse_date(row.get("date")),
                    party_id=party.id,
                    wood_type_id=_get_or_create_lookup(session, WoodType, row.get("wood_type")),
                    location_id=_get_or_create_lookup(session, Location, row.get("location")),
                    vehicle_no=_str(row.get("vehicle")),
                    weight=row.get("weight"),
                    rate=row.get("rate"),
                    freight=row.get("freight") or 0,
                    loading=row.get("loading") or 0,
                    notes=_str(row.get("notes")),
                    created_by=created_by,
                )
            report.created += 1
        except Exception as exc:  # noqa: BLE001
            report.errors.append((i, str(exc)))
    return report
